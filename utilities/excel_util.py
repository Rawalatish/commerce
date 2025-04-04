
import openpyxl
from openpyxl.styles import PatternFill

def get_row_count(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    return (sheet.max_row)

def get_column_count(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    return sheet.max_column

def read_data(file_path, sheet_name, row_num, column_num):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    return sheet.cell(row_num, column_num).value


def write_data(file_path, sheet_name, row_num, column_num,data):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    sheet.cell(row_num, column_num).value = data
    workbook.save(file_path)

def fill_green(file_path, sheet_name, row_num, column_num):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    green_fill = PatternFill(start_color= "00FF00", end_color= "00FF00", patternType= "solid")   # search for hex code for green > 00FF00
    sheet.cell(row_num, column_num).fill = green_fill
    workbook.save(file_path)


def fill_red(file_path, sheet_name, row_num, column_num):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    red_fill = PatternFill(start_color= "FF0000", end_color= "FF0000", patternType= "solid")
    sheet.cell(row_num, column_num).fill = red_fill
    workbook.save(file_path)

